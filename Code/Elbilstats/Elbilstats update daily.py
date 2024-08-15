import pandas as pd
from openpyxl import load_workbook
import requests
from bs4 import BeautifulSoup
from ftplib import FTP
import os
from prefect import flow, task

@flow(name="EVs Registered Daily")
def main_flow():

    # PythonAnywhere credentials
    username = os.getenv('PYTHONANYWHERE_USERNAME')
    api_token = os.getenv('PYTHONANYWHERE_API')
    filename = "top20"
    pythonanywhere_path = f'/home/{username}/mysite/templates/{filename}.xlsx'

    # Download the Excel file from PythonAnywhere
    download_url = f'https://www.pythonanywhere.com/api/v0/user/{username}/files/path{pythonanywhere_path}'
    headers = {'Authorization': f'Token {api_token}'}

    response = requests.get(download_url, headers=headers)
    if response.status_code == 200:
        with open('top20.xlsx', 'wb') as file:
            file.write(response.content)
        print('File downloaded successfully.')
    else:
        print(f'Failed to download the file. Status code: {response.status_code}')
        print(f'Response content: {response.content}')

    # URL of the webpage to scrape
    url = 'https://elbilstatistikk.no/top20/'

    excel_path = "top20.xlsx"
    # Fetch the content of the webpage
    response = requests.get(url)
    response.raise_for_status()  # Raise an exception for HTTP errors

    # Parse the webpage content
    soup = BeautifulSoup(response.text, 'html.parser')

    # Find the relevant section of the webpage
    content_div = soup.select_one('body > div.row.justify-content-center')

    # Extract the data
    data = []
    for row in content_div.find_all('tr')[1:]:  # Skip the header row
        cols = row.find_all('td')
        data.append([col.text.strip() for col in cols])

    # Create a DataFrame
    df = pd.DataFrame(data, columns=['Rank', 'Model', 'Sales'])

    # Write the updated data to the Excel file
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name='Top20', index=False)

    print(f"Data has been scraped and saved to {excel_path}")

    # Define the URL
    url = "https://elbilstatistikk.no/"

    # Send a GET request to the website
    response = requests.get(url)

    # Check if the request was successful
    if response.status_code == 200:
        # Parse the HTML content
        soup = BeautifulSoup(response.content, 'html.parser')

        # Find all tables
        tables = soup.find_all('table')

        # Check if there are at least three tables
        if len(tables) >= 3:
            # Select the third table
            table = tables[2]

            # Extract the table headers
            headers = [header.get_text(strip=True) for header in table.find_all('th')]

            # Extract the table rows
            rows = []
            for row in table.find_all('tr')[1:]:  # Skipping the header row
                cells = row.find_all('td')
                row_data = [cell.get_text(strip=True) for cell in cells]
                rows.append(row_data)

            # Ensure all rows have the same number of columns as the headers
            max_columns = len(headers)
            rows = [row + [''] * (max_columns - len(row)) for row in rows]

            # Create a DataFrame
            df = pd.DataFrame(rows, columns=headers if headers else None)

            # Write the updated data to the Excel file
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                df.to_excel(writer, sheet_name='Today', index=False)
            print("Table saved to elbilstatistikk_table.xlsx")
        else:
            print("Less than three tables found on the webpage.")
    else:
        print(f"Failed to retrieve the content. Status code: {response.status_code}")


    excel_path = "top20.xlsx"

    # Load the workbook and sheets
    workbook = load_workbook(excel_path)
    top20_sheet = workbook["Top20"]
    data_sheet = workbook["Data"]

    # Get the current month
    current_month = pd.Timestamp.now().strftime('%B')
    print(f"Current Month: {current_month}")

    # Read data from B65 to C84 in the Top20 sheet
    top20_data = []
    for row in top20_sheet.iter_rows(min_row=65, max_row=84, min_col=2, max_col=3, values_only=True):
        top20_data.append(row)

    # Create a DataFrame from the extracted data
    df_top20 = pd.DataFrame(top20_data, columns=['Model', 'Sales'])
    print(f"Extracted Data from Top20 Sheet:\n{df_top20}")

    # Find the correct column in the Data sheet for the current month
    month_col = None
    for col in data_sheet.iter_cols(min_row=1, max_row=1, min_col=1, max_col=data_sheet.max_column):
        for cell in col:
            if cell.value == current_month:
                month_col = cell.column
                break
        if month_col:
            break

    # Ensure the month column was found
    if month_col is None:
        raise ValueError(f"Month column for {current_month} not found in the Data sheet.")
    print(f"Month Column: {month_col}")

    # Write the data to the Data sheet starting from the correct cell
    start_row = 3  # Start from row 3
    for idx, row in df_top20.iterrows():
        model_cell = data_sheet.cell(row=start_row + idx, column=month_col)
        sales_cell = data_sheet.cell(row=start_row + idx, column=month_col + 1)
        model_cell.value = row['Model']
        sales_cell.value = row['Sales']
        print(f"Writing to Data Sheet - Row: {start_row + idx}, Model Cell: {model_cell.coordinate}, Sales Cell: {sales_cell.coordinate}")

    # Save the workbook
    workbook.save(excel_path)

    print(f"Data from B65:C84 in Top20 sheet has been copied to {current_month} in Data sheet.")


    def excel_to_json(input_excel_file):
        # Read Excel file
        xls = pd.ExcelFile(input_excel_file)
        
        for sheet_name in xls.sheet_names:
            # Read each sheet
            df = pd.read_excel(input_excel_file, sheet_name=sheet_name)
            
            # # Clean data
            # df = clean_data(df)
            
            # Convert DataFrame to JSON
            json_data = df.to_json(orient='records')
            
            # Write JSON data to file
            output_json_file = f"{sheet_name.lower()}.json"  # Output JSON file name based on sheet name
            with open(output_json_file, 'w') as f:
                f.write(json_data)
                
            # Upload JSON file to PythonAnywhere
            upload_json_to_pythonanywhere(output_json_file)

    def upload_json_to_pythonanywhere(json_file):
        for attempt in range(3):
            # Define PythonAnywhere file path to upload the file to
            username = os.getenv('PYTHONANYWHERE_USERNAME')
            api_token = os.getenv('PYTHONANYWHERE_API')
            pythonanywhere_file_path = f'/home/{username}/mysite/static/assets/json/{json_file}'

            # Define PythonAnywhere upload URL
            upload_url = f'https://www.pythonanywhere.com/api/v0/user/{username}/files/path{pythonanywhere_file_path}'

            # Define headers with API token
            headers = {
                'Authorization': f'Token {api_token}'
            }

            # Open the JSON file
            with open(json_file, 'rb') as file:
                # Create a dictionary with file data
                files = {'content': (json_file, file, 'application/octet-stream')}

                # Make HTTP POST request to upload the file
                response = requests.post(upload_url, headers=headers, files=files)

            # Check if the upload was successful
            if response.status_code == 200:
                print(f'{json_file} - File uploaded successfully to PythonAnywhere.')

                # Upload to FTP server
                ftp_server = 'elbilstats.no'
                ftp_username = 'cvkkwedh'
                ftp_password = '!Adwo20!Adwo20'  # Replace with your actual password
                ftp_directory = 'public_html/static/assets/json'

                upload_file_to_ftp(ftp_server, ftp_username, ftp_password, ftp_directory, json_file)
                break
            else:
                print(f'{json_file}: Error uploading file to PythonAnywhere: {response.text}')
                if attempt == 2:
                    print("Max attempts reached, exiting loop.")

    def upload_file_to_ftp(ftp_server, username, password, directory, file_to_upload):
        for attempt in range(3):
            try:
                # Establish connection to the FTP server
                ftp = FTP(ftp_server)
                ftp.login(user=username, passwd=password)

                # Change to the target directory
                ftp.cwd(directory)

                # Extract file name from file path
                file_name = os.path.basename(file_to_upload)

                # Open the file to upload
                with open(file_to_upload, 'rb') as file:
                    # Upload the file
                    ftp.storbinary(f'STOR {file_name}', file)

                # List files in the directory to verify upload
                ftp.retrlines('LIST')

                # Close the FTP connection
                ftp.quit()
                print(f"File '{file_name}' successfully uploaded to '{directory}' on '{ftp_server}'.")
                break  # Exit the loop if upload is successful

            except Exception as e:
                print(f"Failed to upload file: {e}")
                if attempt == 2:
                    print("Max attempts reached, exiting loop.")

    input_excel_file = 'top20.xlsx'  # Provide the path to your Excel file

    excel_to_json(input_excel_file)

    # PythonAnywhere credentials
    local_file_path = "top20.xlsx"  # Replace with the path to your local file, e.g., "top20.xlsx"
    pythonanywhere_path = f'/home/{username}/mysite/templates/top20.xlsx'  # Replace with the desired path on PythonAnywhere

    # URL for uploading the file
    upload_url = f'https://www.pythonanywhere.com/api/v0/user/{username}/files/path{pythonanywhere_path}'

    # Headers for authentication
    headers = {
        'Authorization': f'Token {api_token}'
    }

    # Function to upload a file to PythonAnywhere
    def upload_file_to_pythonanywhere(local_file_path, upload_url, headers):
        if not os.path.isfile(local_file_path):
            print(f"The file {local_file_path} does not exist.")
            return
        
        try:
            with open(local_file_path, 'rb') as file:
                files = {'content': (os.path.basename(local_file_path), file, 'application/octet-stream')}
                response = requests.post(upload_url, headers=headers, files=files)
            
            if response.status_code == 200:
                print('File uploaded successfully to PythonAnywhere.')
            else:
                print(f'Error uploading file to PythonAnywhere: {response.text}')
                response.raise_for_status()
        except Exception as e:
            print(f'An error occurred: {e}')

    # Call the function to upload the file
    upload_file_to_pythonanywhere(local_file_path, upload_url, headers)

if __name__ == "__main__":
    main_flow()